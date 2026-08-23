from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import inspect
from sqlalchemy.orm import Session

from app.models.inventory_entry import InventoryEntry
from app.models.inventory_session import InventorySession
from app.models.inventory_session_total import InventorySessionTotal
from app.models.item import Item
from app.services.export import (
    ACCOUNTING_SEMIFINISHED_SHEET_TITLE,
    build_xlsx_accounting_template_export,
    is_semifinished_item,
    sort_accounting_export_rows,
)
from app.services.export_repository import (
    fetch_session_catalog_export_rows,
    fetch_session_export_rows,
)

GOODS_SHEET_TITLE = "\u0422\u043e\u0432\u0430\u0440\u044b"
DATA_START_ROW = 8


def _has_table(db: Session, table_name: str) -> bool:
    try:
        return bool(inspect(db.connection()).has_table(table_name))
    except Exception:
        return False


def _string_qty(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _as_sorted_list(values: set[int]) -> list[int]:
    return sorted(int(value) for value in values)


def _is_closed_status(value: object) -> bool:
    raw_value = getattr(value, "value", value)
    normalized = str(raw_value).rsplit(".", 1)[-1].strip().lower()
    return normalized == "closed"


def _footer_start_row(sheet) -> int | None:
    footer_markers = (
        "\u0418\u043d\u0432\u0435\u043d\u0442\u0430\u0440\u0438\u0437\u0430\u0446\u0438\u044e "
        "\u043f\u0440\u043e\u0438\u0437\u0432\u0435\u043b",
        "\u0418\u043d\u0432\u0435\u043d\u0442\u0430\u0440\u0438\u0437\u0430\u0446\u0438\u044e "
        "\u043f\u0440\u0438\u043d\u044f\u043b",
    )
    for row_index in range(sheet.max_row, 0, -1):
        for column_index in range(1, min(sheet.max_column, 5) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if isinstance(value, str) and any(marker in value for marker in footer_markers):
                return row_index
    return None


def _count_written_accounting_rows(sheet) -> int:
    footer_row = _footer_start_row(sheet)
    last_row = (footer_row - 1) if footer_row is not None else sheet.max_row
    count = 0
    for row_index in range(DATA_START_ROW, last_row + 1):
        values = [
            sheet.cell(row=row_index, column=column_index).value for column_index in range(1, 6)
        ]
        if any(value not in (None, "") for value in values):
            count += 1
    return count


def build_session_export_diagnostics(db: Session, session_id: int) -> dict[str, Any]:
    session = db.query(InventorySession).filter(InventorySession.id == session_id).first()
    if session is None:
        return {"session_id": session_id, "exists": False}
    is_closed = bool(getattr(session, "is_closed", False)) or _is_closed_status(session.status)

    entry_rows = (
        db.query(
            InventoryEntry.item_id,
            InventoryEntry.quantity,
            Item.product_code,
            Item.name,
            Item.unit,
            Item.is_active,
        )
        .outerjoin(Item, Item.id == InventoryEntry.item_id)
        .filter(InventoryEntry.session_id == session_id)
        .all()
    )

    totals_table_exists = _has_table(db, InventorySessionTotal.__tablename__)
    if totals_table_exists:
        total_rows = (
            db.query(
                InventorySessionTotal.item_id,
                InventorySessionTotal.qty_final,
                Item.product_code,
                Item.name,
                Item.unit,
                Item.is_active,
            )
            .outerjoin(Item, Item.id == InventorySessionTotal.item_id)
            .filter(InventorySessionTotal.session_id == session_id)
            .all()
        )
    else:
        total_rows = []

    active_catalog_rows = (
        db.query(Item.id, Item.product_code, Item.name, Item.unit, Item.is_active)
        .filter(
            Item.warehouse_id == session.warehouse_id,
            Item.is_active.is_(True),
        )
        .all()
    )

    live_repository_overlay_rows = (
        db.query(InventoryEntry.item_id)
        .join(Item, Item.id == InventoryEntry.item_id)
        .filter(InventoryEntry.session_id == session_id)
        .all()
    )

    _, catalog_rows = fetch_session_catalog_export_rows(db=db, session_id=session_id)
    _, csv_rows = fetch_session_export_rows(db=db, session_id=session_id)

    template_rows: list[dict[str, Any]] = [
        {
            "ItemId": row.item_id,
            "ProductCode": row.product_code,
            "Item": row.name,
            "Unit": row.unit,
            "Qty": row.qty,
            "Category": (str(row.category).strip() or "Uncategorized"),
        }
        for row in catalog_rows
    ]
    regular_rows = [row for row in template_rows if not is_semifinished_item(row)]
    semifinished_rows = [row for row in template_rows if is_semifinished_item(row)]
    sort_accounting_export_rows(regular_rows)
    sort_accounting_export_rows(semifinished_rows)

    workbook_payload = build_xlsx_accounting_template_export(
        [
            {
                "ProductCode": row["ProductCode"],
                "Item": row["Item"],
                "Unit": row["Unit"],
                "Qty": row["Qty"],
                "Category": row["Category"],
            }
            for row in regular_rows
        ],
        extra_sheets=(
            {
                ACCOUNTING_SEMIFINISHED_SHEET_TITLE: [
                    {
                        "ProductCode": row["ProductCode"],
                        "Item": row["Item"],
                        "Unit": row["Unit"],
                        "Qty": row["Qty"],
                        "Category": row["Category"],
                    }
                    for row in semifinished_rows
                ]
            }
            if semifinished_rows
            else None
        ),
    )
    workbook = load_workbook(filename=BytesIO(workbook_payload), data_only=True)
    xlsx_sheet_counts = {
        title: _count_written_accounting_rows(workbook[title])
        for title in (GOODS_SHEET_TITLE, ACCOUNTING_SEMIFINISHED_SHEET_TITLE)
        if title in workbook.sheetnames
    }

    entry_ids = {int(row.item_id) for row in entry_rows}
    total_ids = {int(row.item_id) for row in total_rows}
    active_catalog_ids = {int(row.id) for row in active_catalog_rows}
    live_repository_overlay_ids = {int(row.item_id) for row in live_repository_overlay_rows}
    repository_quantity_ids = (
        total_ids if is_closed and totals_table_exists else live_repository_overlay_ids
    )
    catalog_ids = {int(row.item_id) for row in catalog_rows}
    csv_ids = {int(row.item_id) for row in csv_rows}
    before_split_ids = {int(row["ItemId"]) for row in template_rows}
    regular_ids = {int(row["ItemId"]) for row in regular_rows}
    semifinished_ids = {int(row["ItemId"]) for row in semifinished_rows}

    item_details: dict[int, dict[str, Any]] = {}
    for row in entry_rows:
        item_details.setdefault(
            int(row.item_id),
            {
                "item_id": int(row.item_id),
                "product_code": row.product_code,
                "name": row.name,
                "unit": row.unit,
                "is_active": row.is_active,
            },
        )["entry_qty"] = _string_qty(row.quantity)
    for row in total_rows:
        item_details.setdefault(
            int(row.item_id),
            {
                "item_id": int(row.item_id),
                "product_code": row.product_code,
                "name": row.name,
                "unit": row.unit,
                "is_active": row.is_active,
            },
        )["snapshot_qty"] = _string_qty(row.qty_final)
    for row in catalog_rows:
        item_details.setdefault(
            int(row.item_id),
            {
                "item_id": int(row.item_id),
                "product_code": row.product_code,
                "name": row.name,
                "unit": row.unit,
                "is_active": int(row.item_id) in active_catalog_ids,
            },
        )["export_qty"] = _string_qty(row.qty)

    def details(item_ids: set[int]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for item_id in sorted(item_ids):
            detail = dict(
                item_details.get(
                    item_id,
                    {
                        "item_id": item_id,
                        "product_code": None,
                        "name": None,
                        "unit": None,
                        "is_active": None,
                    },
                )
            )
            detail["in_inventory_entries"] = item_id in entry_ids
            detail["in_inventory_session_totals"] = item_id in total_ids
            detail["in_active_catalog"] = item_id in active_catalog_ids
            detail["in_catalog_export_rows"] = item_id in catalog_ids
            detail["in_csv_export_rows"] = item_id in csv_ids
            rows.append(detail)
        return rows

    transitions = [
        ("inventory_entries", entry_ids, "inventory_session_totals", total_ids),
        ("inventory_session_totals", total_ids, "catalog_export_rows", catalog_ids),
        ("inventory_session_totals", total_ids, "csv_export_rows", csv_ids),
        ("catalog_export_rows", catalog_ids, "before_split", before_split_ids),
        (
            "before_split",
            before_split_ids,
            "regular_plus_semifinished",
            regular_ids | semifinished_ids,
        ),
    ]
    losses = [
        {
            "from": source_name,
            "to": target_name,
            "missing_item_ids": _as_sorted_list(source_ids - target_ids),
            "missing_items": details(source_ids - target_ids),
        }
        for source_name, source_ids, target_name, target_ids in transitions
        if source_ids - target_ids
    ]

    return {
        "session_id": int(session_id),
        "exists": True,
        "status": str(session.status),
        "is_closed": is_closed,
        "warehouse_id": int(session.warehouse_id),
        "counts": {
            "inventory_entries": len(entry_rows),
            "inventory_session_totals": len(total_rows),
            "active_catalog": len(active_catalog_rows),
            "repository_qty_by_item_id": len(repository_quantity_ids),
            "export_repository_catalog_rows": len(catalog_rows),
            "export_repository_csv_rows": len(csv_rows),
            "before_split": len(template_rows),
            "regular_rows": len(regular_rows),
            "semifinished_rows": len(semifinished_rows),
            "xlsx_written_rows": sum(xlsx_sheet_counts.values()),
        },
        "xlsx_written_rows_by_sheet": xlsx_sheet_counts,
        "catalog_gaps": {
            "entry_items_not_in_active_catalog": details(entry_ids - active_catalog_ids),
            "snapshot_items_not_in_active_catalog": details(total_ids - active_catalog_ids),
            "fallback_rows_from_export_source": details(catalog_ids - active_catalog_ids),
        },
        "snapshot_live_gaps": {
            "snapshot_items_without_inventory_entry": details(total_ids - entry_ids),
            "inventory_entries_without_snapshot": details(entry_ids - total_ids),
        },
        "stage_item_ids": {
            "inventory_entries": _as_sorted_list(entry_ids),
            "inventory_session_totals": _as_sorted_list(total_ids),
            "active_catalog": _as_sorted_list(active_catalog_ids),
            "repository_qty_by_item_id": _as_sorted_list(repository_quantity_ids),
            "export_repository_catalog_rows": _as_sorted_list(catalog_ids),
            "export_repository_csv_rows": _as_sorted_list(csv_ids),
            "before_split": _as_sorted_list(before_split_ids),
            "regular_rows": _as_sorted_list(regular_ids),
            "semifinished_rows": _as_sorted_list(semifinished_ids),
        },
        "losses": losses,
    }
