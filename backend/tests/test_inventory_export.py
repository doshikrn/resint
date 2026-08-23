import csv
import time
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.models.inventory_entry import InventoryEntry
from app.models.item import Item
from app.models.user import User
from app.services.accounting_categories import (
    BAKERY,
    CANNED,
    CHEESE_AND_DELI,
    DAIRY,
    DESSERTS,
    FROZEN,
    GROCERY,
    MEAT,
    PRODUCE,
    SEAFOOD,
    SPICES_AND_SAUCES,
    UNCATEGORIZED,
    resolve_accounting_category,
)
from app.services.export import (
    ACCOUNTING_SEMIFINISHED_SHEET_TITLE,
    is_semifinished_item,
    sort_accounting_export_rows,
    sort_export_rows_by_item_name,
)

ACCOUNTING_GROUP_COLUMN = 1
ACCOUNTING_CODE_COLUMN = 2
ACCOUNTING_ITEM_COLUMN = 3
ACCOUNTING_UNIT_COLUMN = 4
ACCOUNTING_QTY_COLUMN = 5


def test_session_export_csv_returns_attachment_with_expected_headers(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("text/csv")

    content_disposition = export.headers.get("content-disposition", "")
    assert content_disposition.startswith('attachment; filename="inventory_')
    assert content_disposition.endswith('_DRAFT.csv"')

    body = export.content.decode("utf-8")
    assert (
        "Zone,Warehouse,SessionId,SessionStatus,Item,Unit,Qty,Category,"
        "CountedOutsideZone,CountedByZone,UpdatedAt,UpdatedBy,Station,Department"
    ) in body
    assert item.name in body


def test_session_export_xlsx_returns_attachment_and_xlsx_payload(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 3, "mode": "set"},
    )
    assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx"},
    )
    assert export.status_code == 200
    assert (
        export.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    content_disposition = export.headers.get("content-disposition", "")
    assert content_disposition.startswith('attachment; filename="inventory_')
    assert content_disposition.endswith('_CLOSED.xlsx"')

    # XLSX is a ZIP container and starts with PK magic bytes.
    assert export.content[:2] == b"PK"


def test_session_export_xlsx_matches_template_spec(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    assert workbook.sheetnames == ["Товары"]

    goods_sheet = workbook["Товары"]
    assert goods_sheet.cell(row=6, column=ACCOUNTING_GROUP_COLUMN).value == "Группа"
    assert goods_sheet.cell(row=8, column=ACCOUNTING_GROUP_COLUMN).value == UNCATEGORIZED
    assert goods_sheet.cell(row=8, column=ACCOUNTING_CODE_COLUMN).value == item.product_code
    assert goods_sheet.cell(row=8, column=ACCOUNTING_ITEM_COLUMN).value == item.name
    assert goods_sheet.cell(row=8, column=ACCOUNTING_UNIT_COLUMN).value in {
        "кг",
        "л",
        "шт",
        item.unit,
    }
    assert isinstance(goods_sheet.cell(row=8, column=ACCOUNTING_QTY_COLUMN).value, (int, float))
    assert goods_sheet.cell(row=8, column=ACCOUNTING_QTY_COLUMN).number_format == "0.###"


def test_sort_export_rows_by_item_name_mixed_case_trim_and_stable_ties():
    rows = [
        {"Item": "  banana "},
        {"Item": "zebra"},
        {"Item": "Apple"},
        {"Item": "aPPle"},
    ]
    sort_export_rows_by_item_name(rows)
    assert [r["Item"] for r in rows] == ["Apple", "aPPle", "  banana ", "zebra"]


def test_sort_accounting_export_rows_groups_categories_and_sorts_items():
    rows = [
        {"Category": "Овощи", "Item": "Яблоко"},
        {"Category": "Бакалея", "Item": "Рис"},
        {"Category": "Бакалея", "Item": "Мука"},
        {"Category": "", "Item": "Без группы"},
    ]

    sort_accounting_export_rows(rows)

    assert [(row["Category"], row["Item"]) for row in rows] == [
        ("Бакалея", "Мука"),
        ("Бакалея", "Рис"),
        ("Овощи", "Яблоко"),
        ("", "Без группы"),
    ]


def test_accounting_category_fallback_prefers_explicit_category_then_reference_code():
    assert (
        resolve_accounting_category(
            "Ручная группа",
            product_code="01564",
            item_name="Маффины",
        )
        == "Ручная группа"
    )
    expected_by_reference_code = {
        "01564": DESSERTS,
        "01313": FROZEN,
        "01082": CANNED,
        "01145": DAIRY,
        "01062": GROCERY,
        "01091": MEAT,
        "00250": PRODUCE,
        "01266": SEAFOOD,
        "01236": SPICES_AND_SAUCES,
        "02700": CHEESE_AND_DELI,
        "01151": BAKERY,
    }
    for product_code, expected_category in expected_by_reference_code.items():
        assert (
            resolve_accounting_category(
                UNCATEGORIZED,
                product_code=product_code,
                item_name="Нейтральное название",
            )
            == expected_category
        )
    assert (
        resolve_accounting_category(
            "",
            product_code="",
            item_name="Неизвестная позиция",
        )
        == UNCATEGORIZED
    )


def test_is_semifinished_item_detects_pf_marker_any_case():
    for label in ("п/ф", "П/ф", "п/Ф", "П/Ф"):
        assert is_semifinished_item({"Item": label})
        assert is_semifinished_item({"Item": f"Заготовка {label} для супа"})
    assert not is_semifinished_item({"Item": "Milk"})
    assert not is_semifinished_item({"Item": "п ф без слэша"})


def _xlsx_goods_rows_on_sheet(content: bytes, sheet_title: str) -> list[tuple[str, str]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    goods_sheet = workbook[sheet_title]
    rows_out: list[tuple[str, str]] = []
    for row_index in range(8, goods_sheet.max_row + 1):
        code = goods_sheet.cell(row=row_index, column=ACCOUNTING_CODE_COLUMN).value
        name = goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        if code and name:
            rows_out.append((str(code), str(name)))
    return rows_out


def test_session_export_xlsx_semifinished_sheet_splits_pf_items(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    milk_name = seed_zone_warehouse_item["item"].name

    payloads = (
        {
            "product_code": "60101",
            "name": "Соус п/ф томат",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "60102",
            "name": "П/ф картофель",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "60103",
            "name": "п/Ф лук",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "60104",
            "name": "П/Ф рис",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "60105",
            "name": "Plain Sugar",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    created_ids: list[int] = []
    for payload in payloads:
        r = client.post("/items", headers=auth_headers, json=payload)
        assert r.status_code == 200
        created_ids.append(r.json()["id"])

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_id in created_ids:
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": item_id, "quantity": 1, "mode": "set"},
        )
        assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    assert "Товары" in workbook.sheetnames
    assert ACCOUNTING_SEMIFINISHED_SHEET_TITLE in workbook.sheetnames
    assert workbook.sheetnames[0] == "Товары"

    main_rows = _xlsx_goods_rows_on_sheet(export.content, "Товары")
    pf_rows = _xlsx_goods_rows_on_sheet(export.content, ACCOUNTING_SEMIFINISHED_SHEET_TITLE)

    main_names = [n for _, n in main_rows]
    pf_names = [n for _, n in pf_rows]

    for name in main_names:
        assert not is_semifinished_item({"Item": name})
    for name in pf_names:
        assert is_semifinished_item({"Item": name})

    assert milk_name in main_names
    assert "Plain Sugar" in main_names
    assert set(pf_names) == {
        "Соус п/ф томат",
        "П/ф картофель",
        "п/Ф лук",
        "П/Ф рис",
    }
    assert main_names == sorted(main_names, key=lambda n: n.strip().lower())
    expected_pf_rows = [
        {"Category": SPICES_AND_SAUCES, "Item": "Соус п/ф томат"},
        {"Category": PRODUCE, "Item": "П/ф картофель"},
        {"Category": PRODUCE, "Item": "п/Ф лук"},
        {"Category": GROCERY, "Item": "П/Ф рис"},
    ]
    sort_accounting_export_rows(expected_pf_rows)
    assert pf_names == [row["Item"] for row in expected_pf_rows]
    assert _xlsx_group_by_item_name_on_sheet(
        export.content,
        ACCOUNTING_SEMIFINISHED_SHEET_TITLE,
    ) == {row["Item"]: row["Category"] for row in expected_pf_rows}

    main_ws = workbook["Товары"]
    pf_ws = workbook[ACCOUNTING_SEMIFINISHED_SHEET_TITLE]
    required_header_merges = {"A6:A7", "B6:C6", "D6:D7", "E6:E7"}
    assert required_header_merges <= {str(cell_range) for cell_range in main_ws.merged_cells.ranges}
    assert required_header_merges <= {str(cell_range) for cell_range in pf_ws.merged_cells.ranges}

    def _footer_marker_row(sheet):
        for row_index in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=1).value
            if isinstance(value, str) and "Инвентаризацию произвел" in value:
                return row_index
        return None

    assert _footer_marker_row(main_ws) is not None
    assert _footer_marker_row(pf_ws) is not None


def test_session_export_xlsx_no_pf_sheet_when_no_semifinished_items(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": seed_zone_warehouse_item["item"].id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200
    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    assert workbook.sheetnames == ["Товары"]


def test_session_export_csv_still_lists_semifinished_items_on_single_export(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    semi = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "60110",
            "name": "Соус П/ф для теста CSV",
            "unit": "l",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert semi.status_code == 200

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": semi.json()["id"], "quantity": 2.5, "mode": "set"},
    )
    assert add.status_code == 200

    export_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert export_csv.status_code == 200
    body = export_csv.content.decode("utf-8")
    assert "Соус П/ф для теста CSV" in body


def test_session_export_xlsx_fallback_semifinished_on_pf_sheet(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    semi = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "60120",
            "name": "Заготовка п/ф inactive",
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert semi.status_code == 200
    semi_id = semi.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": semi_id, "quantity": 7, "mode": "set"},
    )
    assert add.status_code == 200

    item = db_session.query(Item).filter(Item.id == semi_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    assert ACCOUNTING_SEMIFINISHED_SHEET_TITLE in workbook.sheetnames

    pf_rows = _xlsx_goods_rows_on_sheet(export.content, ACCOUNTING_SEMIFINISHED_SHEET_TITLE)
    assert any("Заготовка п/ф inactive" == name for _, name in pf_rows)
    assert all(is_semifinished_item({"Item": name}) for _, name in pf_rows)


def _session_export_csv_item_names(body: str) -> list[str]:
    reader = csv.reader(StringIO(body))
    rows = list(reader)
    if not rows:
        return []
    header = rows[0]
    item_idx = header.index("Item")
    return [r[item_idx] for r in rows[1:] if len(r) > item_idx]


def _session_export_xlsx_item_names(content: bytes) -> list[str]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    goods_sheet = workbook["Товары"]
    names: list[str] = []
    for row_index in range(8, goods_sheet.max_row + 1):
        name = goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        code = goods_sheet.cell(row=row_index, column=ACCOUNTING_CODE_COLUMN).value
        if name and code:
            names.append(str(name))
    return names


def test_session_export_csv_and_xlsx_share_alphabetical_item_order(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    created: dict[str, int] = {}
    for payload in (
        {
            "product_code": "50101",
            "name": "Zulu row",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "50102",
            "name": "alpha mixed",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
        {
            "product_code": "50103",
            "name": "Mike",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    ):
        r = client.post("/items", headers=auth_headers, json=payload)
        assert r.status_code == 200
        created[payload["name"]] = r.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_name in ("Zulu row", "Mike", "alpha mixed"):
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": created[item_name], "quantity": 1, "mode": "set"},
        )
        assert add.status_code == 200

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    subset = frozenset({"Zulu row", "alpha mixed", "Mike"})
    csv_subset = [
        n for n in _session_export_csv_item_names(exp_csv.content.decode("utf-8")) if n in subset
    ]
    xlsx_subset = [n for n in _session_export_xlsx_item_names(exp_xlsx.content) if n in subset]
    expected = sorted(subset, key=lambda n: n.strip().lower())
    assert csv_subset == expected
    assert xlsx_subset == expected
    assert csv_subset == xlsx_subset


def test_session_export_fallback_row_sorted_by_item_name_with_catalog_rows(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    milk_name = seed_zone_warehouse_item["item"].name

    banana = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "50201",
            "name": "Banana line",
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert banana.status_code == 200
    apple = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "50200",
            "name": "Apple gap",
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert apple.status_code == 200
    apple_id = apple.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_id, qty in (
        (seed_zone_warehouse_item["item"].id, 1),
        (banana.json()["id"], 2),
        (apple_id, 3),
    ):
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": item_id, "quantity": qty, "mode": "set"},
        )
        assert add.status_code == 200

    item = db_session.query(Item).filter(Item.id == apple_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    subset = frozenset({"Apple gap", "Banana line", milk_name})
    csv_subset = [
        n for n in _session_export_csv_item_names(exp_csv.content.decode("utf-8")) if n in subset
    ]
    xlsx_subset = [n for n in _session_export_xlsx_item_names(exp_xlsx.content) if n in subset]
    expected = sorted(subset, key=lambda n: n.strip().lower())
    assert csv_subset == expected
    assert xlsx_subset == expected
    assert csv_subset == xlsx_subset


def test_session_export_unknown_session_returns_404(client, auth_headers):
    export = client.get(
        "/inventory/sessions/999999/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert export.status_code == 404


def test_session_export_forbidden_for_cook_on_foreign_session(
    client,
    auth_headers,
    auth_headers_cook,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers_cook,
        params={"format": "csv"},
    )

    assert export.status_code == 403


def test_session_export_allowed_for_chef_on_foreign_session(
    client,
    auth_headers,
    auth_headers_chef,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers_chef,
        params={"format": "csv"},
    )

    assert export.status_code == 200


def test_session_export_allowed_for_souschef_on_own_session(
    client,
    auth_headers_souschef,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers_souschef,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers_souschef,
        params={"format": "csv"},
    )

    assert export.status_code == 200


def test_session_export_accepts_accounting_v1_template(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200


def test_session_export_rejects_unknown_template(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv", "template": "accounting_v2"},
    )
    assert export.status_code == 422


def test_session_export_entries_sorted_and_qty_preserved_and_uncategorized(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    category = client.post(
        "/items/categories",
        headers=auth_headers,
        json={"name": "Meat"},
    )
    assert category.status_code == 201
    category_id = category.json()["id"]

    item_with_category = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "10200",
            "name": "Beef Round",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
            "category_id": category_id,
        },
    )
    assert item_with_category.status_code == 200
    item_with_category_id = item_with_category.json()["id"]

    item_without_category = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "10201",
            "name": "Water Bottle",
            "unit": "l",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert item_without_category.status_code == 200
    item_without_category_id = item_without_category.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add_first = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_with_category_id, "quantity": 1.13, "mode": "set"},
    )
    assert add_first.status_code == 200

    add_second = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_without_category_id, "quantity": 2.37, "mode": "set"},
    )
    assert add_second.status_code == 200

    patch_first_step = client.patch(
        f"/items/{item_with_category_id}",
        headers=auth_headers,
        json={"step": 0.25},
    )
    assert patch_first_step.status_code == 200

    patch_second_step = client.patch(
        f"/items/{item_without_category_id}",
        headers=auth_headers,
        json={"step": 0.5},
    )
    assert patch_second_step.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]
    # Trailing empty rows should be trimmed; max_row should be close to data rows
    assert goods_sheet.max_row >= 8  # at least header + some data rows

    rows = []
    for row_index in range(8, goods_sheet.max_row + 1):
        code = goods_sheet.cell(row=row_index, column=ACCOUNTING_CODE_COLUMN).value
        name = goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        unit = goods_sheet.cell(row=row_index, column=ACCOUNTING_UNIT_COLUMN).value
        qty = goods_sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value
        if code and name:
            rows.append((str(code), str(name), str(unit), qty))

    row_by_name = {row[1]: row for row in rows}
    assert row_by_name["Beef Round"][0] == "10200"
    assert row_by_name["Water Bottle"][0] == "10201"
    assert row_by_name["Beef Round"][3] == 1.13
    assert row_by_name["Water Bottle"][3] == 2.37
    assert [row[1] for row in rows] == ["Beef Round", "Milk", "Water Bottle"]


def test_export_xlsx_keeps_fractional_precision_for_qty_915(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    create_item = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "10915",
            "name": "Precision Test Item",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert create_item.status_code == 200
    item_id = create_item.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 9.15, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]

    qty_by_item_name = {
        str(goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value): goods_sheet.cell(
            row=row_index, column=ACCOUNTING_QTY_COLUMN
        ).value
        for row_index in range(8, goods_sheet.max_row + 1)
        if goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
    }

    assert qty_by_item_name["Precision Test Item"] == 9.15


def test_export_draft_session_download_has_template_sheet(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200
    assert export.headers.get("content-disposition", "").startswith("attachment; filename=")

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    assert workbook.sheetnames == ["Товары"]


def test_export_xlsx_preserves_accounting_footer_block(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]

    footer_rows = []
    for row_index in range(1, goods_sheet.max_row + 1):
        cell_value = goods_sheet.cell(row=row_index, column=1).value
        if isinstance(cell_value, str) and "Инвентаризацию произвел" in cell_value:
            footer_rows.append((row_index, cell_value))

    assert footer_rows, "Accounting footer block must be preserved in exported XLSX"
    assert footer_rows[0][0] >= 9
    assert footer_rows[0][0] <= 12


def test_export_closed_session_filename_has_closed_suffix(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    content_disposition = export.headers.get("content-disposition", "")
    assert content_disposition.endswith('_CLOSED.xlsx"')


def test_export_xlsx_qty_is_numeric_for_excel_sum(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]
    assert isinstance(goods_sheet.cell(row=8, column=ACCOUNTING_QTY_COLUMN).value, (int, float))


def test_export_xlsx_includes_all_catalog_items_and_dash_for_missing_qty(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    measured = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "12001",
            "name": "Measured Item",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert measured.status_code == 200
    measured_id = measured.json()["id"]

    unmeasured = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "12002",
            "name": "Unmeasured Item",
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert unmeasured.status_code == 200

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": measured_id, "quantity": 9.15, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]

    values_by_name = {}
    for row_index in range(8, goods_sheet.max_row + 1):
        item_name = goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        if item_name:
            values_by_name[str(item_name)] = {
                "code": goods_sheet.cell(row=row_index, column=ACCOUNTING_CODE_COLUMN).value,
                "unit": goods_sheet.cell(row=row_index, column=ACCOUNTING_UNIT_COLUMN).value,
                "qty": goods_sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value,
            }

    assert "Measured Item" in values_by_name
    assert "Unmeasured Item" in values_by_name
    assert values_by_name["Measured Item"]["qty"] == 9.15
    assert values_by_name["Unmeasured Item"]["qty"] == "-"
    assert values_by_name["Measured Item"]["code"] == "12001"
    assert values_by_name["Unmeasured Item"]["code"] == "12002"


def test_export_xlsx_keeps_session_item_even_if_item_is_inactive(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    item_response = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "12003",
            "name": "Лист лайма",
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
        },
    )
    assert item_response.status_code == 200
    item_id = item_response.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 4, "mode": "set"},
    )
    assert add.status_code == 200

    item = db_session.query(Item).filter(Item.id == item_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    goods_sheet = workbook["Товары"]

    values_by_name = {}
    for row_index in range(8, goods_sheet.max_row + 1):
        item_name = goods_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        if item_name:
            values_by_name[str(item_name)] = {
                "code": goods_sheet.cell(row=row_index, column=ACCOUNTING_CODE_COLUMN).value,
                "unit": goods_sheet.cell(row=row_index, column=ACCOUNTING_UNIT_COLUMN).value,
                "qty": goods_sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value,
            }

    assert "Лист лайма" in values_by_name
    assert values_by_name["Лист лайма"]["code"] == "12003"
    assert values_by_name["Лист лайма"]["qty"] == 4


def test_export_csv_keeps_russian_names_utf8(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    create_item = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "10202",
            "name": "Говядина вырезка",
            "unit": "кг",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert create_item.status_code == 200
    item_id = create_item.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 2.5, "mode": "set"},
    )
    assert add.status_code == 200

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    body = export.content.decode("utf-8")
    assert "Говядина вырезка" in body


def _csv_qty_by_item_name(body: str, item_name: str) -> float | None:
    reader = csv.DictReader(StringIO(body))
    for row in reader:
        if row.get("Item") == item_name:
            return float(row["Qty"])
    return None


def _xlsx_qty_by_item_name_on_sheet(
    content: bytes, sheet_name: str, item_name: str
) -> float | str | None:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook[sheet_name]
    for row_index in range(8, sheet.max_row + 1):
        if str(sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value or "") == item_name:
            qty = sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value
            if qty == "-":
                return None
            return qty
    return None


def _xlsx_qty_by_item_name(content: bytes, item_name: str) -> float | str | None:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_index in range(8, sheet.max_row + 1):
            if (
                str(sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value or "")
                == item_name
            ):
                qty = sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value
                if qty == "-":
                    return None
                return qty
    return None


def _csv_counted_qty_by_item_name(body: str) -> dict[str, float]:
    reader = csv.DictReader(StringIO(body))
    return {str(row["Item"]): float(row["Qty"]) for row in reader if row.get("Item")}


def _csv_category_by_item_name(body: str, item_name: str) -> str | None:
    reader = csv.DictReader(StringIO(body))
    for row in reader:
        if row.get("Item") == item_name:
            return str(row.get("Category") or "")
    return None


def _xlsx_group_by_item_name_on_sheet(content: bytes, sheet_name: str) -> dict[str, str]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook[sheet_name]
    group_by_item_name: dict[str, str] = {}
    current_group = ""
    for row_index in range(8, sheet.max_row + 1):
        group_value = sheet.cell(row=row_index, column=ACCOUNTING_GROUP_COLUMN).value
        if group_value not in (None, ""):
            current_group = str(group_value)
        item_name = sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        if item_name:
            group_by_item_name[str(item_name)] = current_group
    return group_by_item_name


def _xlsx_counted_qty_by_item_name(content: bytes) -> dict[str, float]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    quantities: dict[str, float] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_index in range(8, sheet.max_row + 1):
            item_name = sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
            qty = sheet.cell(row=row_index, column=ACCOUNTING_QTY_COLUMN).value
            if not item_name or qty in (None, "-"):
                continue
            quantities[str(item_name)] = float(qty)
    return quantities


def test_session_export_xlsx_adds_group_column_and_merges_category_block(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    category_name = "Овощи, зелень, фрукты"
    category = client.post(
        "/items/categories",
        headers=auth_headers,
        json={"name": category_name},
    )
    assert category.status_code == 201

    item_names = ("Абрикос для группы", "Картофель для группы")
    for product_code, item_name in zip(("70101", "70102"), item_names, strict=True):
        response = client.post(
            "/items",
            headers=auth_headers,
            json={
                "product_code": product_code,
                "name": item_name,
                "unit": "kg",
                "warehouse_id": warehouse.id,
                "step": 0.01,
                "category_id": category.json()["id"],
            },
        )
        assert response.status_code == 200

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    export = client.get(
        f"/inventory/sessions/{active.json()['id']}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    workbook = load_workbook(filename=BytesIO(export.content), data_only=True)
    sheet = workbook["Товары"]
    assert sheet["A6"].value == "Группа"
    assert sheet["B7"].value == "Код"
    assert sheet["C7"].value == "Наименование"
    assert sheet["D6"].value == "Ед. изм."
    assert sheet["E6"].value == "Остаток фактический"
    assert str(sheet.print_area).endswith("$A$1:$E$305")

    group_by_item = _xlsx_group_by_item_name_on_sheet(export.content, "Товары")
    assert {group_by_item[item_name] for item_name in item_names} == {category_name}

    item_rows = [
        row_index
        for row_index in range(8, sheet.max_row + 1)
        if sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value in item_names
    ]
    assert len(item_rows) == 2
    assert item_rows[1] == item_rows[0] + 1
    assert f"A{item_rows[0]}:A{item_rows[1]}" in {
        str(cell_range) for cell_range in sheet.merged_cells.ranges
    }
    assert [
        sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value for row_index in item_rows
    ] == list(item_names)


def test_closed_session_uncategorized_pf_items_use_accounting_groups_in_csv_and_xlsx(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    expected_groups = {
        "Говядина мякоть отварная П/Ф": MEAT,
        "Куриное филе в сливочном соусе П/Ф": MEAT,
        "Демиглас соус П/Ф": SPICES_AND_SAUCES,
        "Лосось жареный П/Ф": SEAFOOD,
        "Картофельное пюре П/Ф": PRODUCE,
        "Жареный рис ланч П/Ф": GROCERY,
        "Крамбл П/Ф": DESSERTS,
        "Пассировка на суп ланч П/Ф": PRODUCE,
    }

    item_ids: list[int] = []
    for item_name in expected_groups:
        response = client.post(
            "/items",
            headers=auth_headers,
            json={
                "name": item_name,
                "unit": "kg",
                "warehouse_id": warehouse.id,
                "step": 0.01,
            },
        )
        assert response.status_code == 200
        assert response.json()["category_id"] is None
        item_ids.append(response.json()["id"])

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_id in item_ids:
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": item_id, "quantity": 1.25, "mode": "set"},
        )
        assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    xlsx_groups = _xlsx_group_by_item_name_on_sheet(
        exp_xlsx.content,
        ACCOUNTING_SEMIFINISHED_SHEET_TITLE,
    )
    csv_groups = {
        item_name: _csv_category_by_item_name(
            exp_csv.content.decode("utf-8"),
            item_name,
        )
        for item_name in expected_groups
    }
    assert {item_name: xlsx_groups[item_name] for item_name in expected_groups} == expected_groups
    assert csv_groups == expected_groups
    assert UNCATEGORIZED not in xlsx_groups.values()

    workbook = load_workbook(filename=BytesIO(exp_xlsx.content), data_only=True)
    pf_sheet = workbook[ACCOUNTING_SEMIFINISHED_SHEET_TITLE]
    meat_rows = [
        row_index
        for row_index in range(8, pf_sheet.max_row + 1)
        if pf_sheet.cell(row=row_index, column=ACCOUNTING_ITEM_COLUMN).value
        in {
            "Говядина мякоть отварная П/Ф",
            "Куриное филе в сливочном соусе П/Ф",
        }
    ]
    assert len(meat_rows) == 2
    assert meat_rows[1] == meat_rows[0] + 1
    assert f"A{meat_rows[0]}:A{meat_rows[1]}" in {
        str(cell_range) for cell_range in pf_sheet.merged_cells.ranges
    }


def test_closed_session_export_xlsx_catalog_matches_csv_snapshot_qty(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    """XLSX catalog path must use the same closed-session qty source as CSV (totals snapshot)."""
    warehouse = seed_zone_warehouse_item["warehouse"]
    item = seed_zone_warehouse_item["item"]
    milk_name = item.name

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item.id, "quantity": 2.25, "mode": "set"},
    )
    assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    correct = client.patch(
        f"/inventory/sessions/{session_id}/entries/{item.id}",
        headers={**auth_headers, "If-Match": "1"},
        json={"quantity": 9.99, "reason": "post-close recount"},
    )
    assert correct.status_code == 200

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    csv_qty = _csv_qty_by_item_name(exp_csv.content.decode("utf-8"), milk_name)
    xlsx_qty = _xlsx_qty_by_item_name_on_sheet(exp_xlsx.content, "Товары", milk_name)
    assert csv_qty == 2.25
    assert isinstance(xlsx_qty, (int, float))
    assert float(xlsx_qty) == 2.25


def test_export_qty_preserved_after_pf_sheet_split(
    client,
    auth_headers,
    seed_zone_warehouse_item,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    semi = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "70101",
            "name": "Тестовая п/ф смесь",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert semi.status_code == 200
    regular = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "70102",
            "name": "Zebra Regular",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert regular.status_code == 200

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_id, qty in (
        (semi.json()["id"], 3.33),
        (regular.json()["id"], 7.77),
    ):
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": item_id, "quantity": qty, "mode": "set"},
        )
        assert add.status_code == 200

    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_xlsx.status_code == 200

    assert (
        float(_xlsx_qty_by_item_name_on_sheet(exp_xlsx.content, "Товары", "Zebra Regular")) == 7.77
    )
    assert (
        float(
            _xlsx_qty_by_item_name_on_sheet(
                exp_xlsx.content, ACCOUNTING_SEMIFINISHED_SHEET_TITLE, "Тестовая п/ф смесь"
            )
        )
        == 3.33
    )

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert exp_csv.status_code == 200
    body = exp_csv.content.decode("utf-8")
    assert _csv_qty_by_item_name(body, "Zebra Regular") == 7.77
    assert _csv_qty_by_item_name(body, "Тестовая п/ф смесь") == 3.33


def test_closed_session_inactive_item_survives_csv_and_xlsx_export_after_close(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    item_response = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80100",
            "name": "Closed Inactive Snapshot",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert item_response.status_code == 200
    item_id = item_response.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 8.75, "mode": "set"},
    )
    assert add.status_code == 200
    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    item = db_session.query(Item).filter(Item.id == item_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    assert (
        _csv_qty_by_item_name(exp_csv.content.decode("utf-8"), "Closed Inactive Snapshot") == 8.75
    )
    assert _xlsx_qty_by_item_name(exp_xlsx.content, "Closed Inactive Snapshot") == 8.75


def test_export_diagnostics_keeps_snapshot_item_after_live_entry_deleted(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]

    item_response = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80101",
            "name": "Forensic Lost Row",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert item_response.status_code == 200
    item_id = item_response.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 6.5, "mode": "set"},
    )
    assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    item = db_session.query(Item).filter(Item.id == item_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    delete = client.delete(
        f"/inventory/sessions/{session_id}/entries/{item_id}",
        headers=auth_headers,
    )
    assert delete.status_code == 204

    report = client.get(f"/inventory/reports/session/{session_id}", headers=auth_headers)
    assert report.status_code == 200
    report_items = {row["item_id"]: row for row in report.json()["items"]}
    assert report_items[item_id]["quantity"] == 6.5

    diagnostic = client.get(
        f"/inventory/sessions/{session_id}/export/diagnostics",
        headers=auth_headers,
    )
    assert diagnostic.status_code == 200
    body = diagnostic.json()

    assert body["counts"]["inventory_entries"] == 0
    assert body["counts"]["inventory_session_totals"] == 1
    assert body["counts"]["export_repository_csv_rows"] == 1
    assert item_id in body["stage_item_ids"]["export_repository_catalog_rows"]
    assert item_id in body["stage_item_ids"]["export_repository_csv_rows"]
    assert (
        body["snapshot_live_gaps"]["snapshot_items_without_inventory_entry"][0]["item_id"]
        == item_id
    )
    assert not any(
        loss["from"] == "inventory_session_totals"
        and loss["to"] == "catalog_export_rows"
        and item_id in loss["missing_item_ids"]
        for loss in body["losses"]
    )


def test_closed_session_xlsx_keeps_snapshot_item_after_live_entry_deleted(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item_response = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80102",
            "name": "Forensic Snapshot Only",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert item_response.status_code == 200
    item_id = item_response.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 4.25, "mode": "set"},
    )
    assert add.status_code == 200
    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    item = db_session.query(Item).filter(Item.id == item_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    delete = client.delete(
        f"/inventory/sessions/{session_id}/entries/{item_id}",
        headers=auth_headers,
    )
    assert delete.status_code == 204

    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert export.status_code == 200

    names = _session_export_xlsx_item_names(export.content)
    assert "Forensic Snapshot Only" in names
    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert exp_csv.status_code == 200
    assert _csv_qty_by_item_name(exp_csv.content.decode("utf-8"), "Forensic Snapshot Only") == 4.25
    assert (
        _xlsx_qty_by_item_name_on_sheet(export.content, "Товары", "Forensic Snapshot Only") == 4.25
    )


def test_closed_session_csv_and_xlsx_counted_snapshot_rows_remain_identical(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    pf_name = "Snapshot parity \u043f/\u0444 sauce"

    regular = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80103",
            "name": "Snapshot parity regular",
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    pf = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80104",
            "name": pf_name,
            "unit": "kg",
            "warehouse_id": warehouse.id,
            "step": 0.01,
        },
    )
    assert regular.status_code == 200
    assert pf.status_code == 200

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    for item_id, qty in (
        (regular.json()["id"], 5.5),
        (pf.json()["id"], 1.25),
    ):
        add = client.post(
            f"/inventory/sessions/{session_id}/entries",
            headers=auth_headers,
            json={"item_id": item_id, "quantity": qty, "mode": "set"},
        )
        assert add.status_code == 200

    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    for item_id in (regular.json()["id"], pf.json()["id"]):
        item = db_session.query(Item).filter(Item.id == item_id).first()
        assert item is not None
        item.is_active = False
        db_session.add(item)
    db_session.commit()

    for item_id in (regular.json()["id"], pf.json()["id"]):
        delete = client.delete(
            f"/inventory/sessions/{session_id}/entries/{item_id}",
            headers=auth_headers,
        )
        assert delete.status_code == 204

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_csv.status_code == 200
    assert exp_xlsx.status_code == 200

    expected = {
        "Snapshot parity regular": 5.5,
        pf_name: 1.25,
    }
    csv_quantities = _csv_counted_qty_by_item_name(exp_csv.content.decode("utf-8"))
    xlsx_quantities = _xlsx_counted_qty_by_item_name(exp_xlsx.content)
    assert {name: csv_quantities[name] for name in expected} == expected
    assert {name: xlsx_quantities[name] for name in expected} == expected


def test_closed_session_pf_snapshot_item_exported_to_semifinished_sheet(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    item_name = "Deleted entry \u043f/\u0444 snapshot"
    category_name = "Соусы и полуфабрикаты"
    category = client.post(
        "/items/categories",
        headers=auth_headers,
        json={"name": category_name},
    )
    assert category.status_code == 201

    item_response = client.post(
        "/items",
        headers=auth_headers,
        json={
            "product_code": "80105",
            "name": item_name,
            "unit": "pcs",
            "warehouse_id": warehouse.id,
            "step": 1.0,
            "category_id": category.json()["id"],
        },
    )
    assert item_response.status_code == 200
    item_id = item_response.json()["id"]

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    add = client.post(
        f"/inventory/sessions/{session_id}/entries",
        headers=auth_headers,
        json={"item_id": item_id, "quantity": 9, "mode": "set"},
    )
    assert add.status_code == 200
    close = client.post(f"/inventory/sessions/{session_id}/close", headers=auth_headers)
    assert close.status_code == 200

    item = db_session.query(Item).filter(Item.id == item_id).first()
    assert item is not None
    item.is_active = False
    db_session.add(item)
    db_session.commit()

    delete = client.delete(
        f"/inventory/sessions/{session_id}/entries/{item_id}",
        headers=auth_headers,
    )
    assert delete.status_code == 204

    exp_xlsx = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    assert exp_xlsx.status_code == 200
    workbook = load_workbook(filename=BytesIO(exp_xlsx.content), data_only=True)
    assert ACCOUNTING_SEMIFINISHED_SHEET_TITLE in workbook.sheetnames
    assert (
        _xlsx_qty_by_item_name_on_sheet(
            exp_xlsx.content, ACCOUNTING_SEMIFINISHED_SHEET_TITLE, item_name
        )
        == 9
    )
    assert (
        _xlsx_group_by_item_name_on_sheet(exp_xlsx.content, ACCOUNTING_SEMIFINISHED_SHEET_TITLE)[
            item_name
        ]
        == category_name
    )

    exp_csv = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "csv"},
    )
    assert exp_csv.status_code == 200
    assert _csv_category_by_item_name(exp_csv.content.decode("utf-8"), item_name) == category_name


def test_export_500_rows_completes_quickly(
    client,
    auth_headers,
    seed_zone_warehouse_item,
    db_session,
):
    warehouse = seed_zone_warehouse_item["warehouse"]
    actor = db_session.query(User).filter(User.username == "testuser").first()
    assert actor is not None

    active = client.post(
        "/inventory/sessions/active",
        headers=auth_headers,
        json={"warehouse_id": warehouse.id},
    )
    assert active.status_code == 200
    session_id = active.json()["id"]

    items = [
        Item(
            product_code=f"{10000 + index:05d}",
            name=f"Perf Item {index:03d}",
            unit="pcs",
            step=1.0,
            warehouse_id=warehouse.id,
            is_active=True,
        )
        for index in range(1, 501)
    ]
    db_session.add_all(items)
    db_session.flush()

    entries = [
        InventoryEntry(
            session_id=session_id,
            item_id=item.id,
            quantity=float(index),
            updated_by_user_id=actor.id,
        )
        for index, item in enumerate(items, start=1)
    ]
    db_session.add_all(entries)
    db_session.commit()

    started = time.perf_counter()
    export = client.get(
        f"/inventory/sessions/{session_id}/export",
        headers=auth_headers,
        params={"format": "xlsx", "template": "accounting_v1"},
    )
    elapsed_seconds = time.perf_counter() - started

    assert export.status_code == 200
    assert elapsed_seconds < 8.0
