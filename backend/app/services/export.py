from __future__ import annotations

import csv
import re
from copy import copy
from collections.abc import Iterable
from datetime import UTC, datetime, timezone, timedelta
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.print_settings import PrintArea

ALMATY_TZ = timezone(timedelta(hours=6))

ENTRY_COLUMNS = [
    "ProductCode",
    "Zone",
    "Warehouse",
    "SessionId",
    "SessionStatus",
    "Item",
    "Unit",
    "Qty",
    "Category",
    "CountedOutsideZone",
    "CountedByZone",
    "UpdatedAt",
    "UpdatedBy",
    "Station",
    "Department",
]

ACCOUNTING_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "templates" / "accounting_v1.xlsx"
)


def _safe_slug(value: str) -> str:
    normalized = re.sub(r"\s+", "_", value.strip().lower())
    normalized = re.sub(r"[^a-z0-9_\-]+", "", normalized)
    return normalized or "warehouse"


def build_export_filename(
    warehouse_name: str, session_created_at: datetime, status: str, file_ext: str
) -> str:
    date_part = session_created_at.date().isoformat()
    warehouse_part = _safe_slug(warehouse_name)
    status_part = _safe_slug(status).upper()
    return f"inventory_{warehouse_part}_{date_part}_{status_part}.{file_ext}"


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(ALMATY_TZ).replace(tzinfo=None)


def _qty_number_format(unit: str) -> str:
    normalized = (unit or "").strip().lower()
    if normalized in {"kg", "l", "кг", "л"}:
        return "0.00"
    if normalized in {"pcs", "шт"}:
        return "0"
    return "0.00"


def sort_export_rows_by_item_name(rows: list[dict]) -> None:
    """Sort session export row dicts by ``Item`` before CSV/XLSX builders.

    Comparison uses trimmed names and case-insensitive ordering. The sort is
    stable so ties keep their relative order.
    """
    rows.sort(key=lambda r: str(r.get("Item", "") or "").strip().lower())


def _accounting_category_label(value: object) -> str:
    return str(value or "").strip() or "Uncategorized"


def sort_accounting_export_rows(rows: list[dict]) -> None:
    """Group accounting rows by category and keep item names alphabetical inside each group."""

    def sort_key(row: dict) -> tuple[bool, str, str]:
        category = _accounting_category_label(row.get("Category"))
        item_name = str(row.get("Item", "") or "").strip()
        return (
            category.casefold() == "uncategorized",
            category.casefold(),
            item_name.casefold(),
        )

    rows.sort(key=sort_key)


_SEMIFINISHED_MARKER_CF = "п/ф".casefold()

# Excel worksheet titles cannot contain ``/``. Use Unicode FRACTION SLASH (U+2044), same look as "п/ф".
ACCOUNTING_SEMIFINISHED_SHEET_TITLE = "п\u2044ф"


def is_semifinished_item(row: dict) -> bool:
    """True if accounting export row name (``Item``) contains semi-finished marker п/ф (any case)."""
    item = str(row.get("Item", "") or "")
    return _SEMIFINISHED_MARKER_CF in item.casefold()


def build_csv_export(rows: Iterable[dict]) -> bytes:
    buffer = StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(ENTRY_COLUMNS)
    for row in rows:
        updated_at = row.get("UpdatedAt")
        writer.writerow(
            [
                row.get("ProductCode", ""),
                row.get("Zone", ""),
                row.get("Warehouse", ""),
                row.get("SessionId", ""),
                row.get("SessionStatus", ""),
                row.get("Item", ""),
                _unit_label_ru(row.get("Unit", "")),
                row.get("Qty", ""),
                row.get("Category", ""),
                row.get("CountedOutsideZone", ""),
                row.get("CountedByZone", ""),
                updated_at.isoformat() if isinstance(updated_at, datetime) else "",
                row.get("UpdatedBy", ""),
                row.get("Station", ""),
                row.get("Department", ""),
            ]
        )
    return buffer.getvalue().encode("utf-8")


def build_xlsx_export(
    rows: Iterable[dict],
    summary: dict,
) -> bytes:
    workbook = Workbook()

    entries_sheet = workbook.active
    entries_sheet.title = "Entries"
    entries_sheet.append(ENTRY_COLUMNS)
    for index, column_name in enumerate(ENTRY_COLUMNS, start=1):
        entries_sheet.cell(row=1, column=index).font = Font(bold=True)

    for row in rows:
        entries_sheet.append(
            [
                row.get("ProductCode", ""),
                row.get("Zone", ""),
                row.get("Warehouse", ""),
                int(row.get("SessionId", 0)),
                row.get("SessionStatus", ""),
                row.get("Item", ""),
                _unit_label_ru(row.get("Unit", "")),
                row.get("Qty", 0),
                row.get("Category", ""),
                row.get("CountedOutsideZone", ""),
                row.get("CountedByZone", ""),
                _excel_datetime(row.get("UpdatedAt")),
                row.get("UpdatedBy", ""),
                row.get("Station", ""),
                row.get("Department", ""),
            ]
        )

    entries_sheet.freeze_panes = "A2"

    for data_row in range(2, entries_sheet.max_row + 1):
        unit_value = str(entries_sheet.cell(row=data_row, column=7).value or "")
        qty_cell = entries_sheet.cell(row=data_row, column=8)
        qty_cell.number_format = _qty_number_format(unit_value)

        updated_at_cell = entries_sheet.cell(row=data_row, column=12)
        if updated_at_cell.value is not None:
            updated_at_cell.number_format = "yyyy-mm-dd hh:mm:ss"

    item_col_width = max(20, len("Item") + 2)
    for data_row in range(2, entries_sheet.max_row + 1):
        item_value = str(entries_sheet.cell(row=data_row, column=6).value or "")
        item_col_width = max(item_col_width, len(item_value) + 2)
    entries_sheet.column_dimensions["F"].width = item_col_width

    summary_sheet = workbook.create_sheet(title="Summary")
    summary_rows = [
        ("ReportVersion", summary.get("ReportVersion", "v1")),
        ("GeneratedAt", _excel_datetime(summary.get("GeneratedAt"))),
        ("Zone", summary.get("Zone", "")),
        ("Warehouse", summary.get("Warehouse", "")),
        ("SessionId", summary.get("SessionId", "")),
        ("SessionStatus", summary.get("SessionStatus", "")),
        ("SessionStartedAt", _excel_datetime(summary.get("SessionStartedAt"))),
        ("SessionClosedAt", _excel_datetime(summary.get("SessionClosedAt"))),
        ("TotalLines", summary.get("TotalLines", 0)),
    ]

    for key, value in summary_rows:
        summary_sheet.append([key, value])

    for row_index in (2, 7, 8):
        value_cell = summary_sheet.cell(row=row_index, column=2)
        if value_cell.value is not None:
            value_cell.number_format = "yyyy-mm-dd hh:mm:ss"

    unit_start_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=unit_start_row, column=1, value="TotalQtyByUnit")
    summary_sheet.cell(row=unit_start_row, column=1).font = Font(bold=True)
    summary_sheet.cell(row=unit_start_row + 1, column=1, value="Unit")
    summary_sheet.cell(row=unit_start_row + 1, column=2, value="SumQty")
    summary_sheet.cell(row=unit_start_row + 1, column=1).font = Font(bold=True)
    summary_sheet.cell(row=unit_start_row + 1, column=2).font = Font(bold=True)

    row_ptr = unit_start_row + 2
    for unit, qty in sorted(
        (summary.get("TotalQtyByUnit") or {}).items(), key=lambda pair: pair[0]
    ):
        summary_sheet.cell(row=row_ptr, column=1, value=_unit_label_ru(unit))
        cell = summary_sheet.cell(row=row_ptr, column=2, value=qty)
        cell.number_format = _qty_number_format(str(unit))
        row_ptr += 1

    totals_by_category = summary.get("TotalsByCategory") or {}
    if totals_by_category:
        category_start_row = row_ptr + 1
        summary_sheet.cell(row=category_start_row, column=1, value="TotalsByCategory")
        summary_sheet.cell(row=category_start_row, column=1).font = Font(bold=True)
        summary_sheet.cell(row=category_start_row + 1, column=1, value="Category")
        summary_sheet.cell(row=category_start_row + 1, column=2, value="Lines")
        summary_sheet.cell(row=category_start_row + 1, column=3, value="SumQty")
        summary_sheet.cell(row=category_start_row + 1, column=1).font = Font(bold=True)
        summary_sheet.cell(row=category_start_row + 1, column=2).font = Font(bold=True)
        summary_sheet.cell(row=category_start_row + 1, column=3).font = Font(bold=True)

        category_row = category_start_row + 2
        for category_name, stats in sorted(totals_by_category.items(), key=lambda pair: pair[0]):
            summary_sheet.cell(row=category_row, column=1, value=category_name)
            summary_sheet.cell(row=category_row, column=2, value=int(stats.get("lines", 0)))
            summary_sheet.cell(row=category_row, column=3, value=stats.get("sum_qty", 0))
            category_row += 1

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _unit_label_ru(unit: str) -> str:
    normalized = (unit or "").strip().lower()
    if normalized == "kg":
        return "кг"
    if normalized == "l":
        return "л"
    if normalized == "pcs":
        return "шт"
    return unit


def _find_footer_start_row(goods_sheet) -> int | None:
    footer_markers = (
        "Инвентаризацию произвел",
        "Инвентаризацию принял",
    )
    for row_index in range(goods_sheet.max_row, 0, -1):
        for column_index in range(1, min(goods_sheet.max_column, 5) + 1):
            value = goods_sheet.cell(row=row_index, column=column_index).value
            if not isinstance(value, str):
                continue
            if any(marker in value for marker in footer_markers):
                return row_index
    return None


def _clear_accounting_goods_data_values(
    sheet,
    *,
    data_start_row: int,
    last_row: int,
    last_column: int = 5,
) -> None:
    """Clear product table cell values while preserving the accounting layout."""
    if last_row < data_start_row:
        return
    _unmerge_accounting_group_data_cells(sheet, data_start_row=data_start_row, last_row=last_row)
    for row_index in range(data_start_row, last_row + 1):
        for col_index in range(1, last_column + 1):
            sheet.cell(row=row_index, column=col_index).value = None


def _unmerge_accounting_group_data_cells(
    sheet,
    *,
    data_start_row: int,
    last_row: int | None = None,
) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        is_group_data_range = (
            merged_range.min_col == 1
            and merged_range.max_col == 1
            and merged_range.min_row >= data_start_row
        )
        if not is_group_data_range:
            continue
        if last_row is not None and merged_range.min_row > last_row:
            continue
        sheet.unmerge_cells(str(merged_range))


def _merge_accounting_group_cells(
    sheet,
    rows: list[dict],
    *,
    data_start_row: int,
) -> None:
    group_start = 0
    while group_start < len(rows):
        category = _accounting_category_label(rows[group_start].get("Category"))
        group_end = group_start
        while group_end + 1 < len(rows):
            next_category = _accounting_category_label(rows[group_end + 1].get("Category"))
            if next_category.casefold() != category.casefold():
                break
            group_end += 1

        first_excel_row = data_start_row + group_start
        last_excel_row = data_start_row + group_end
        group_cell = sheet.cell(row=first_excel_row, column=1, value=category)
        group_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        if last_excel_row > first_excel_row:
            sheet.merge_cells(
                start_row=first_excel_row,
                start_column=1,
                end_row=last_excel_row,
                end_column=1,
            )

        group_start = group_end + 1


def _write_accounting_goods_data_rows(sheet, rows: list[dict], *, data_start_row: int = 8) -> int:
    """Write accounting-style columns for ``rows``. Returns last used data row (may be ``data_start_row - 1``)."""
    footer_start_row = _find_footer_start_row(sheet)
    _unmerge_accounting_group_data_cells(
        sheet,
        data_start_row=data_start_row,
        last_row=(footer_start_row - 1) if footer_start_row is not None else None,
    )

    for index, row in enumerate(rows):
        excel_row = data_start_row + index
        sheet.cell(
            row=excel_row,
            column=1,
            value=_accounting_category_label(row.get("Category")),
        )
        sheet.cell(row=excel_row, column=2, value=str(row.get("ProductCode", "")))
        sheet.cell(row=excel_row, column=3, value=str(row.get("Item", "")))
        sheet.cell(row=excel_row, column=4, value=_unit_label_ru(str(row.get("Unit", ""))))

        qty = row.get("Qty")
        qty_cell = sheet.cell(row=excel_row, column=5)
        if qty is None:
            qty_cell.value = "-"
        else:
            qty_cell.value = qty
            qty_cell.number_format = "0.###"

    if not rows:
        return data_start_row - 1
    _merge_accounting_group_cells(sheet, rows, data_start_row=data_start_row)
    return data_start_row + len(rows) - 1


def _trim_accounting_goods_table_area(sheet, *, data_start_row: int, last_data_row: int) -> None:
    """Remove placeholder rows between data and footer (template) or sheet end."""
    footer_start_row = _find_footer_start_row(sheet)
    if footer_start_row is not None:
        if footer_start_row > last_data_row + 1:
            trailing_count = footer_start_row - last_data_row - 1
            if trailing_count > 0:
                footer_merges = _pop_merged_ranges_at_or_after(sheet, start_row=footer_start_row)
                sheet.delete_rows(last_data_row + 1, trailing_count)
                _restore_shifted_merged_ranges(sheet, footer_merges, row_offset=-trailing_count)
        return

    total_rows = sheet.max_row
    trailing_count = total_rows - last_data_row
    if trailing_count > 0:
        sheet.delete_rows(last_data_row + 1, trailing_count)


def _pop_merged_ranges_at_or_after(sheet, *, start_row: int) -> list[tuple[int, int, int, int]]:
    coordinates: list[tuple[int, int, int, int]] = []
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row < start_row:
            continue
        coordinates.append(
            (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
        )
        sheet.unmerge_cells(str(merged_range))
    return coordinates


def _restore_shifted_merged_ranges(
    sheet,
    ranges: list[tuple[int, int, int, int]],
    *,
    row_offset: int,
) -> None:
    for min_row, min_col, max_row, max_col in ranges:
        sheet.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col,
            end_row=max_row + row_offset,
            end_column=max_col,
        )


def _extend_accounting_template_merges(sheet) -> None:
    footer_start_row = _find_footer_start_row(sheet)
    for merged_range in list(sheet.merged_cells.ranges):
        should_extend = merged_range.max_col == 4 and (
            merged_range.max_row <= 5
            or (
                footer_start_row is not None
                and merged_range.min_row <= footer_start_row <= merged_range.max_row
            )
        )
        if not should_extend:
            continue
        old_coordinate = str(merged_range)
        new_coordinate = (
            f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:"
            f"E{merged_range.max_row}"
        )
        sheet.unmerge_cells(old_coordinate)
        sheet.merge_cells(new_coordinate)


def _extend_accounting_print_area(sheet) -> None:
    print_area = sheet.print_area
    if not print_area:
        return
    parsed_print_area = PrintArea.from_string(str(print_area))
    ranges = list(parsed_print_area.ranges)
    if not ranges:
        return
    updated_ranges = []
    for cell_range in ranges:
        max_col = 5 if cell_range.max_col == 4 else cell_range.max_col
        updated_ranges.append(
            f"{get_column_letter(cell_range.min_col)}{cell_range.min_row}:"
            f"{get_column_letter(max_col)}{cell_range.max_row}"
        )
    sheet.print_area = ",".join(updated_ranges)


def _upgrade_legacy_accounting_goods_layout(sheet, *, data_start_row: int = 8) -> None:
    footer_start_row = _find_footer_start_row(sheet)
    old_widths = {column: sheet.column_dimensions[column].width for column in ("A", "B", "C", "D")}
    header_styles = {
        "product_left": copy(sheet["A6"]._style),
        "product_right": copy(sheet["B6"]._style),
        "code": copy(sheet["A7"]._style),
        "name": copy(sheet["B7"]._style),
        "unit_top": copy(sheet["C6"]._style),
        "unit_bottom": copy(sheet["C7"]._style),
        "qty_top": copy(sheet["D6"]._style),
        "qty_bottom": copy(sheet["D7"]._style),
    }

    _extend_accounting_template_merges(sheet)
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= 6 and merged_range.max_row <= 7:
            sheet.unmerge_cells(str(merged_range))

    for row_index in (6, 7):
        for column_index in range(1, 6):
            sheet.cell(row=row_index, column=column_index).value = None

    header_cells = {
        "A6": ("Группа", header_styles["unit_top"]),
        "A7": (None, header_styles["unit_bottom"]),
        "B6": ("Товар", header_styles["product_left"]),
        "C6": (None, header_styles["product_right"]),
        "B7": ("Код", header_styles["code"]),
        "C7": ("Наименование", header_styles["name"]),
        "D6": ("Ед. изм.", header_styles["unit_top"]),
        "D7": (None, header_styles["unit_bottom"]),
        "E6": ("Остаток фактический", header_styles["qty_top"]),
        "E7": (None, header_styles["qty_bottom"]),
    }
    for coordinate, (value, style) in header_cells.items():
        cell = sheet[coordinate]
        cell.value = value
        cell._style = copy(style)

    for coordinate in ("A6:A7", "B6:C6", "D6:D7", "E6:E7"):
        sheet.merge_cells(coordinate)

    last_template_data_row = footer_start_row - 1 if footer_start_row is not None else sheet.max_row
    for row_index in range(data_start_row, last_template_data_row + 1):
        code_style = copy(sheet.cell(row=row_index, column=1)._style)
        name_style = copy(sheet.cell(row=row_index, column=2)._style)
        unit_style = copy(sheet.cell(row=row_index, column=3)._style)
        qty_style = copy(sheet.cell(row=row_index, column=4)._style)
        sheet.cell(row=row_index, column=1)._style = copy(name_style)
        sheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        sheet.cell(row=row_index, column=2)._style = code_style
        sheet.cell(row=row_index, column=3)._style = name_style
        sheet.cell(row=row_index, column=4)._style = unit_style
        sheet.cell(row=row_index, column=5)._style = qty_style

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = old_widths["A"]
    sheet.column_dimensions["C"].width = old_widths["B"]
    sheet.column_dimensions["D"].width = old_widths["C"]
    sheet.column_dimensions["E"].width = old_widths["D"]
    _extend_accounting_print_area(sheet)


def _ensure_accounting_goods_header_row(sheet, *, header_row: int = 7) -> None:
    top_row = header_row - 1
    sheet.merge_cells(
        start_row=top_row,
        start_column=1,
        end_row=header_row,
        end_column=1,
    )
    sheet.merge_cells(
        start_row=top_row,
        start_column=2,
        end_row=top_row,
        end_column=3,
    )
    sheet.merge_cells(
        start_row=top_row,
        start_column=4,
        end_row=header_row,
        end_column=4,
    )
    sheet.merge_cells(
        start_row=top_row,
        start_column=5,
        end_row=header_row,
        end_column=5,
    )
    for row, column, value in (
        (top_row, 1, "Группа"),
        (top_row, 2, "Товар"),
        (header_row, 2, "Код"),
        (header_row, 3, "Наименование"),
        (top_row, 4, "Ед. изм."),
        (top_row, 5, "Остаток фактический"),
    ):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ensure_accounting_goods_group_layout(sheet) -> None:
    if str(sheet["A6"].value or "").strip() == "Группа":
        return
    is_legacy_layout = (
        str(sheet["A6"].value or "").strip() == "Товар"
        and str(sheet["A7"].value or "").strip() == "Код"
        and str(sheet["B7"].value or "").strip() == "Наименование"
    )
    if is_legacy_layout:
        _upgrade_legacy_accounting_goods_layout(sheet)
        return
    _ensure_accounting_goods_header_row(sheet)


def build_xlsx_accounting_template_export(
    rows: Iterable[dict],
    *,
    extra_sheets: dict[str, list[dict]] | None = None,
) -> bytes:
    if ACCOUNTING_TEMPLATE_PATH.exists():
        workbook = load_workbook(filename=ACCOUNTING_TEMPLATE_PATH)
        if "Товары" not in workbook.sheetnames:
            raise ValueError("Template sheet 'Товары' not found")
        goods_sheet = workbook["Товары"]
    else:
        workbook = Workbook()
        goods_sheet = workbook.active
        goods_sheet.title = "Товары"

    _ensure_accounting_goods_group_layout(goods_sheet)

    data_start_row = 8
    normalized_rows = list(rows)
    last_main_row = _write_accounting_goods_data_rows(
        goods_sheet, normalized_rows, data_start_row=data_start_row
    )
    _trim_accounting_goods_table_area(
        goods_sheet, data_start_row=data_start_row, last_data_row=last_main_row
    )

    for sheet_title, extra_rows in (extra_sheets or {}).items():
        if not extra_rows:
            continue
        extra_list = list(extra_rows)
        m_extra = len(extra_list)
        n_main = len(normalized_rows)

        if sheet_title in workbook.sheetnames:
            workbook.remove(workbook[sheet_title])

        pf_sheet = workbook.copy_worksheet(goods_sheet)
        pf_sheet.title = sheet_title

        footer_row = _find_footer_start_row(pf_sheet)
        if footer_row is not None:
            capacity = footer_row - data_start_row
            if m_extra > capacity:
                inserted_row_count = m_extra - capacity
                footer_merges = _pop_merged_ranges_at_or_after(pf_sheet, start_row=footer_row)
                pf_sheet.insert_rows(footer_row, amount=inserted_row_count)
                _restore_shifted_merged_ranges(
                    pf_sheet, footer_merges, row_offset=inserted_row_count
                )

        if n_main > 0:
            _clear_accounting_goods_data_values(
                pf_sheet,
                data_start_row=data_start_row,
                last_row=data_start_row + n_main - 1,
            )

        last_extra = _write_accounting_goods_data_rows(
            pf_sheet, extra_list, data_start_row=data_start_row
        )
        _trim_accounting_goods_table_area(
            pf_sheet, data_start_row=data_start_row, last_data_row=last_extra
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
