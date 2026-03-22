import csv
import re
from datetime import UTC, datetime, timedelta
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy import and_, case, func, literal, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.deps import get_current_user
from app.core.roles import can_manage_catalog
from app.db.session import get_db
from app.models.inventory_entry import InventoryEntry
from app.models.inventory_entry_event import InventoryEntryEvent
from app.models.inventory_session import InventorySession
from app.models.item import Item
from app.models.item_alias import ItemAlias
from app.models.item_category import ItemCategory
from app.models.item_usage_stat import ItemUsageStat
from app.models.station import Station
from app.models.user import User
from app.models.warehouse import Warehouse
from app.schemas.item import (
    ALLOWED_UNITS,
    UNIT_LABELS,
    ItemAliasCreate,
    ItemAliasOut,
    ItemCategoryCreate,
    ItemCategoryOut,
    ItemCreate,
    ItemOut,
    ItemPatch,
    ItemsBulkUpsertRequest,
    ItemUnitOut,
    normalize_name_for_dedupe,
)

router = APIRouter(prefix="/items", tags=["items"])

_PERIOD_RE = re.compile(r"^(\d+)([dwm])$")
_TRUE_VALUES = {"1", "true", "yes", "y", "да"}
_FALSE_VALUES = {"0", "false", "no", "n", "нет"}


from app.core.clock import utc_now as _utc_now


def _parse_period(period: str) -> timedelta:
    raw = period.strip().lower()
    match = _PERIOD_RE.fullmatch(raw)
    if not match:
        raise HTTPException(status_code=422, detail="period must match format like 30d, 2w, 1m")

    value = int(match.group(1))
    unit = match.group(2)
    if value <= 0:
        raise HTTPException(status_code=422, detail="period value must be greater than 0")

    if unit == "d":
        return timedelta(days=value)
    if unit == "w":
        return timedelta(weeks=value)
    return timedelta(days=value * 30)


def _ensure_warehouse_exists(db: Session, warehouse_id: int) -> None:
    exists = db.query(Warehouse.id).filter(Warehouse.id == warehouse_id).first()
    if not exists:
        raise HTTPException(status_code=404, detail="Warehouse not found")


def _resolve_user_warehouse_id(current_user: User) -> int | None:
    warehouse_id = getattr(current_user, "warehouse_id", None)
    if warehouse_id is None:
        warehouse_id = getattr(current_user, "default_warehouse_id", None)
    if warehouse_id is None:
        return None
    return int(warehouse_id)


def _require_warehouse_access(warehouse_id: int, current_user: User) -> None:
    user_warehouse_id = _resolve_user_warehouse_id(current_user)
    if user_warehouse_id is None:
        raise HTTPException(status_code=403, detail="User is not bound to a warehouse")
    if int(warehouse_id) != user_warehouse_id:
        raise HTTPException(status_code=403, detail="Forbidden for this warehouse")


def _require_catalog_manage_role(current_user: User) -> None:
    if not can_manage_catalog(current_user.role):
        raise HTTPException(status_code=403, detail="Only souschef or chef can manage catalog")


def _ensure_category_exists(db: Session, category_id: int | None) -> None:
    if category_id is None:
        return
    exists = db.query(ItemCategory.id).filter(ItemCategory.id == category_id).first()
    if not exists:
        raise HTTPException(status_code=404, detail="Category not found")


def _ensure_station_exists(db: Session, station_id: int | None) -> None:
    if station_id is None:
        return
    exists = db.query(Station.id).filter(Station.id == station_id).first()
    if not exists:
        raise HTTPException(status_code=404, detail="Station not found")


def _to_optional_int(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    return int(text)


def _to_optional_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (float, int)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    return float(text)


def _to_optional_bool(value) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in _TRUE_VALUES:
        return True
    if text in _FALSE_VALUES:
        return False
    raise ValueError("must be boolean")


def _read_tabular_rows(filename: str | None, raw: bytes) -> list[dict[str, object]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        decoded = None
        for encoding in ("utf-8-sig", "cp1251"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise HTTPException(status_code=422, detail="CSV encoding must be utf-8 or cp1251")

        reader = csv.DictReader(StringIO(decoded))
        if not reader.fieldnames:
            raise HTTPException(status_code=422, detail="CSV must contain header row")
        rows: list[dict[str, object]] = []
        for row in reader:
            rows.append({(key or "").strip().lower(): value for key, value in row.items()})
        return rows

    if lower_name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=422, detail="XLSX import requires openpyxl") from exc

        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=422, detail="XLSX is empty")

        headers = [(str(cell).strip().lower() if cell is not None else "") for cell in all_rows[0]]
        if not any(headers):
            raise HTTPException(status_code=422, detail="XLSX must contain header row")

        parsed_rows: list[dict[str, object]] = []
        for source_row in all_rows[1:]:
            parsed = {}
            for idx, value in enumerate(source_row):
                header = headers[idx] if idx < len(headers) else ""
                if header:
                    parsed[header] = value
            parsed_rows.append(parsed)
        return parsed_rows

    raise HTTPException(status_code=422, detail="Only .csv and .xlsx files are supported")


def _build_export_rows(
    db: Session,
    warehouse_id: int | None,
    category_id: int | None,
) -> list[tuple[Item, str | None, str | None]]:
    query = (
        db.query(Item, ItemCategory.name.label("category_name"), Station.name.label("station_name"))
        .outerjoin(ItemCategory, Item.category_id == ItemCategory.id)
        .outerjoin(Station, Item.station_id == Station.id)
    )
    if warehouse_id is not None:
        query = query.filter(Item.warehouse_id == warehouse_id)
    if category_id is not None:
        query = query.filter(Item.category_id == category_id)
    return query.order_by(Item.warehouse_id.asc(), Item.product_code.asc(), Item.name.asc()).all()


@router.get("", response_model=list[ItemOut])
def list_items(
    warehouse_id: int | None = None,
    category_id: int | None = None,
    q: str | None = Query(default=None, min_length=1),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = db.query(Item).filter(Item.is_active == True)  # noqa: E712
    if warehouse_id:
        query = query.filter(Item.warehouse_id == warehouse_id)
    if category_id is not None:
        query = query.filter(Item.category_id == category_id)
    if q:
        query = query.filter(or_(Item.name.ilike(f"%{q}%"), Item.product_code.ilike(f"%{q}%")))
    return query.order_by(Item.product_code.asc(), Item.name.asc()).all()


@router.get("/units", response_model=list[ItemUnitOut])
def list_item_units(_=Depends(get_current_user)):
    return [{"code": unit, "label": UNIT_LABELS[unit]} for unit in ALLOWED_UNITS]


@router.get("/recent", response_model=list[ItemOut])
def recent_items(
    warehouse_id: int,
    session_id: int | None = None,
    limit: int = Query(20, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_warehouse_exists(db, warehouse_id)
    _require_warehouse_access(warehouse_id, current_user)

    if session_id is not None:
        session_exists = (
            db.query(InventorySession.id)
            .filter(
                InventorySession.id == session_id,
                InventorySession.warehouse_id == warehouse_id,
            )
            .first()
        )
        if not session_exists:
            raise HTTPException(status_code=404, detail="Session not found")

    recent_subquery = (
        db.query(
            InventoryEntry.item_id.label("item_id"),
            func.max(InventoryEntry.updated_at).label("last_used_at"),
        )
        .join(InventorySession, InventoryEntry.session_id == InventorySession.id)
        .filter(
            InventorySession.warehouse_id == warehouse_id,
            InventoryEntry.updated_by_user_id == current_user.id,
        )
    )

    if session_id is not None:
        recent_subquery = recent_subquery.filter(InventoryEntry.session_id == session_id)

    recent_subquery = recent_subquery.group_by(InventoryEntry.item_id).subquery()

    return (
        db.query(Item)
        .join(recent_subquery, recent_subquery.c.item_id == Item.id)
        .filter(Item.is_active == True, Item.warehouse_id == warehouse_id)  # noqa: E712
        .order_by(recent_subquery.c.last_used_at.desc(), Item.name.asc())
        .limit(limit)
        .all()
    )


@router.get("/frequent", response_model=list[ItemOut])
def frequent_items(
    warehouse_id: int,
    session_id: int | None = None,
    limit: int = Query(20, ge=1, le=50),
    period: str = Query("30d"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_warehouse_exists(db, warehouse_id)
    _require_warehouse_access(warehouse_id, current_user)
    cutoff = _utc_now() - _parse_period(period)

    if session_id is not None:
        session_exists = (
            db.query(InventorySession.id)
            .filter(
                InventorySession.id == session_id,
                InventorySession.warehouse_id == warehouse_id,
            )
            .first()
        )
        if not session_exists:
            raise HTTPException(status_code=404, detail="Session not found")

    usage_query = (
        db.query(
            InventoryEntryEvent.item_id.label("item_id"),
            func.count(InventoryEntryEvent.id).label("use_count"),
            func.max(InventoryEntryEvent.created_at).label("last_used_at"),
        )
        .join(InventorySession, InventoryEntryEvent.session_id == InventorySession.id)
        .filter(
            InventorySession.warehouse_id == warehouse_id,
            InventoryEntryEvent.actor_user_id == current_user.id,
            InventoryEntryEvent.created_at >= cutoff,
        )
    )

    if session_id is not None:
        usage_query = usage_query.filter(InventoryEntryEvent.session_id == session_id)

    usage_subquery = usage_query.group_by(InventoryEntryEvent.item_id).subquery()

    return (
        db.query(Item)
        .join(usage_subquery, usage_subquery.c.item_id == Item.id)
        .filter(
            Item.is_active == True,  # noqa: E712
            Item.warehouse_id == warehouse_id,
        )
        .order_by(
            usage_subquery.c.use_count.desc(), usage_subquery.c.last_used_at.desc(), Item.name.asc()
        )
        .limit(limit)
        .all()
    )


@router.get("/search", response_model=list[ItemOut])
def search_items(
    q: str = Query(..., min_length=1, max_length=80),
    warehouse_id: int | None = None,
    zone_id: int | None = None,
    category_id: int | None = None,
    limit: int = Query(30, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _ = current_user

    q_clean = q.strip()
    if not q_clean:
        return []

    query = db.query(Item).filter(Item.is_active == True)  # noqa: E712
    if warehouse_id is not None:
        query = query.filter(Item.warehouse_id == warehouse_id)
    elif zone_id is not None:
        query = query.join(Warehouse, Item.warehouse_id == Warehouse.id).filter(
            Warehouse.zone_id == zone_id
        )
    if category_id is not None:
        query = query.filter(Item.category_id == category_id)

    starts = f"{q_clean}%"
    contains = f"%{q_clean}%"

    alias_starts_exists = (
        db.query(literal(1))
        .filter(
            ItemAlias.item_id == Item.id,
            ItemAlias.alias_text.ilike(starts),
        )
        .exists()
    )
    alias_contains_exists = (
        db.query(literal(1))
        .filter(
            ItemAlias.item_id == Item.id,
            ItemAlias.alias_text.ilike(contains),
        )
        .exists()
    )

    query = query.filter(
        or_(
            Item.name.ilike(contains),
            Item.product_code.ilike(contains),
            alias_contains_exists,
        )
    )

    priority = case(
        (Item.product_code.ilike(starts), 0),
        (Item.name.ilike(starts), 1),
        (alias_starts_exists, 2),
        else_=3,
    )
    query = query.order_by(priority.asc(), Item.product_code.asc(), Item.name.asc())

    return query.limit(limit).all()


@router.post("", response_model=ItemOut)
def create_item(payload: ItemCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not can_manage_catalog(current_user.role):
        raise HTTPException(status_code=403, detail="Insufficient role to create items")
    wh = db.query(Warehouse).filter(Warehouse.id == payload.warehouse_id).first()
    if not wh:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    _ensure_category_exists(db, payload.category_id)
    _ensure_station_exists(db, payload.station_id)

    if payload.product_code is not None:
        duplicate = (
            db.query(Item.id)
            .filter(func.lower(Item.product_code) == payload.product_code.lower())
            .first()
        )
        if duplicate:
            raise HTTPException(
                status_code=409, detail="Item with this product_code already exists"
            )

    item = Item(
        product_code=payload.product_code,
        name=payload.name,
        unit=payload.unit,
        step=payload.step,
        min_qty=payload.min_qty,
        max_qty=payload.max_qty,
        is_favorite=payload.is_favorite,
        warehouse_id=payload.warehouse_id,
        category_id=payload.category_id,
        station_id=payload.station_id,
    )
    db.add(item)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Item with this product_code already exists")
    db.refresh(item)
    return item


@router.post("/import")
async def import_items(
    file: UploadFile = File(...),
    dry_run: bool = True,
    default_warehouse_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="Uploaded file is empty")

    rows = _read_tabular_rows(file.filename, raw)
    if not rows:
        return {
            "dry_run": dry_run,
            "total": 0,
            "created": 0,
            "updated": 0,
            "errors": [],
        }

    created = 0
    updated = 0
    errors: list[dict[str, object]] = []

    category_by_name: dict[str, int] = {
        entry.name.lower(): entry.id for entry in db.query(ItemCategory).all()
    }
    station_by_name: dict[str, int] = {
        entry.name.lower(): entry.id for entry in db.query(Station).all()
    }

    for row_index, row in enumerate(rows, start=2):
        try:
            name = str(row.get("name") or "").strip()
            product_code = str(
                row.get("product_code") or row.get("code") or row.get("item_code") or ""
            ).strip()
            unit = str(row.get("unit") or "").strip()
            raw_wh = row.get("warehouse_id", default_warehouse_id)
            warehouse_id = _to_optional_int(raw_wh)
            if warehouse_id is None:
                raise ValueError("warehouse_id is required")

            raw_category_id = row.get("category_id")
            category_id = _to_optional_int(raw_category_id)
            category_name = str(row.get("category") or row.get("category_name") or "").strip()
            if category_id is None and category_name:
                category_id = category_by_name.get(category_name.lower())
                if category_id is None:
                    raise ValueError(f"unknown category '{category_name}'")

            raw_station_id = row.get("station_id")
            station_id = _to_optional_int(raw_station_id)
            station_name = str(row.get("station") or row.get("station_name") or "").strip()
            if station_id is None and station_name:
                station_id = station_by_name.get(station_name.lower())
                if station_id is None:
                    raise ValueError(f"unknown station '{station_name}'")

            _ensure_warehouse_exists(db, warehouse_id)
            _ensure_category_exists(db, category_id)
            _ensure_station_exists(db, station_id)

            step = _to_optional_float(row.get("step"))
            min_qty = _to_optional_float(row.get("min_qty"))
            max_qty = _to_optional_float(row.get("max_qty"))
            is_active = _to_optional_bool(row.get("is_active"))
            is_favorite = _to_optional_bool(row.get("is_favorite"))

            payload = ItemCreate(
                product_code=product_code,
                name=name,
                unit=unit,
                warehouse_id=warehouse_id,
                step=step if step is not None else 1.0,
                min_qty=min_qty,
                max_qty=max_qty,
                is_favorite=is_favorite if is_favorite is not None else False,
                category_id=category_id,
                station_id=station_id,
            )

            existing = None
            if payload.product_code is not None:
                existing = (
                    db.query(Item)
                    .filter(Item.warehouse_id == payload.warehouse_id)
                    .filter(func.lower(Item.product_code) == payload.product_code.lower())
                    .first()
                )

            if existing is None:
                norm_name = normalize_name_for_dedupe(payload.name)
                existing = (
                    db.query(Item)
                    .filter(Item.warehouse_id == payload.warehouse_id)
                    .filter(func.lower(func.trim(Item.name)) == norm_name)
                    .filter(func.lower(Item.unit) == payload.unit.strip().lower())
                    .first()
                )

            if existing:
                updated += 1
                if not dry_run:
                    existing.unit = payload.unit
                    existing.product_code = payload.product_code
                    existing.step = payload.step
                    existing.min_qty = payload.min_qty
                    existing.max_qty = payload.max_qty
                    existing.category_id = payload.category_id
                    existing.station_id = payload.station_id
                    existing.is_favorite = payload.is_favorite
                    if is_active is not None:
                        existing.is_active = is_active
            else:
                created += 1
                if not dry_run:
                    db.add(
                        Item(
                            product_code=payload.product_code,
                            name=payload.name,
                            unit=payload.unit,
                            step=payload.step,
                            min_qty=payload.min_qty,
                            max_qty=payload.max_qty,
                            is_favorite=payload.is_favorite,
                            is_active=is_active if is_active is not None else True,
                            warehouse_id=payload.warehouse_id,
                            category_id=payload.category_id,
                            station_id=payload.station_id,
                        )
                    )
        except Exception as exc:
            errors.append({"row": row_index, "message": str(exc)})

    if not dry_run and not errors:
        db.commit()
    else:
        db.rollback()

    return {
        "dry_run": dry_run,
        "total": len(rows),
        "created": created,
        "updated": updated,
        "errors": errors,
    }


@router.get("/export")
def export_items(
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    warehouse_id: int | None = None,
    category_id: int | None = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    if warehouse_id is not None:
        _ensure_warehouse_exists(db, warehouse_id)
    if category_id is not None:
        _ensure_category_exists(db, category_id)

    rows = _build_export_rows(db=db, warehouse_id=warehouse_id, category_id=category_id)
    header = [
        "id",
        "product_code",
        "name",
        "unit",
        "warehouse_id",
        "category_id",
        "category_name",
        "station_id",
        "station_name",
        "step",
        "min_qty",
        "max_qty",
        "is_active",
        "is_favorite",
    ]

    if format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(header)
        for item, category_name, station_name in rows:
            writer.writerow(
                [
                    item.id,
                    item.product_code,
                    item.name,
                    item.unit,
                    item.warehouse_id,
                    item.category_id,
                    category_name,
                    item.station_id,
                    station_name,
                    item.step,
                    item.min_qty,
                    item.max_qty,
                    item.is_active,
                    item.is_favorite,
                ]
            )
        return Response(
            content=buffer.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="items_export.csv"'},
        )

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=422, detail="XLSX export requires openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(header)
    for item, category_name, station_name in rows:
        ws.append(
            [
                item.id,
                item.product_code,
                item.name,
                item.unit,
                item.warehouse_id,
                item.category_id,
                category_name,
                item.station_id,
                station_name,
                item.step,
                item.min_qty,
                item.max_qty,
                item.is_active,
                item.is_favorite,
            ]
        )

    output = BytesIO()
    wb.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="items_export.xlsx"'},
    )


@router.patch("/{item_id}", response_model=ItemOut)
def patch_item(
    item_id: int,
    payload: ItemPatch,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    data = payload.model_dump(exclude_unset=True)
    if not data:
        return item

    _ensure_category_exists(db, data.get("category_id"))
    _ensure_station_exists(db, data.get("station_id"))

    product_code = data.get("product_code")
    if product_code is not None:
        duplicate = (
            db.query(Item.id)
            .filter(Item.id != item.id, func.lower(Item.product_code) == product_code.lower())
            .first()
        )
        if duplicate:
            raise HTTPException(
                status_code=409, detail="Item with this product_code already exists"
            )

    next_min = data.get("min_qty", item.min_qty)
    next_max = data.get("max_qty", item.max_qty)
    if next_min is not None and next_max is not None and next_min > next_max:
        raise HTTPException(status_code=422, detail="min_qty cannot be greater than max_qty")

    for field, value in data.items():
        setattr(item, field, value)

    db.commit()
    db.refresh(item)
    return item


@router.post("/bulk-upsert")
def bulk_upsert_items(
    payload: ItemsBulkUpsertRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    if not payload.rows:
        return {
            "dry_run": payload.dry_run,
            "total": 0,
            "created": 0,
            "updated": 0,
            "skipped_existing": 0,
            "errors": [],
        }

    created = 0
    updated = 0
    skipped_existing = 0
    errors: list[dict[str, object]] = []

    existing_code_set: set[str] = {
        code.lower() for (code,) in db.query(Item.product_code).all() if code
    }
    existing_name_unit_set: set[tuple[int, str, str]] = {
        (warehouse_id, normalize_name_for_dedupe(name), unit.strip().lower())
        for warehouse_id, name, unit in db.query(
            Item.warehouse_id, Item.name, Item.unit
        ).all()
        if name
    }

    category_by_name: dict[str, int] = {
        entry.name.lower(): entry.id for entry in db.query(ItemCategory).all()
    }
    station_by_name: dict[str, int] = {
        entry.name.lower(): entry.id for entry in db.query(Station).all()
    }

    for row_index, row in enumerate(payload.rows, start=1):
        try:
            warehouse_id = (
                row.warehouse_id if row.warehouse_id is not None else payload.default_warehouse_id
            )
            if warehouse_id is None:
                raise ValueError("warehouse_id is required")

            category_id = row.category_id
            if category_id is None and row.category_name:
                category_id = category_by_name.get(row.category_name.strip().lower())
                if category_id is None:
                    raise ValueError(f"unknown category '{row.category_name}'")

            station_id = row.station_id
            if station_id is None and row.station_name:
                station_id = station_by_name.get(row.station_name.strip().lower())
                if station_id is None:
                    raise ValueError(f"unknown station '{row.station_name}'")

            _ensure_warehouse_exists(db, warehouse_id)
            _ensure_category_exists(db, category_id)
            _ensure_station_exists(db, station_id)

            item_payload = ItemCreate(
                product_code=row.product_code,
                name=row.name,
                unit=row.unit,
                warehouse_id=warehouse_id,
                step=row.step if row.step is not None else 1.0,
                min_qty=row.min_qty,
                max_qty=row.max_qty,
                is_favorite=row.is_favorite if row.is_favorite is not None else False,
                category_id=category_id,
                station_id=station_id,
            )

            existing = None
            if item_payload.product_code is not None:
                existing = (
                    db.query(Item)
                    .filter(Item.warehouse_id == item_payload.warehouse_id)
                    .filter(func.lower(Item.product_code) == item_payload.product_code.lower())
                    .first()
                )

            if existing is None:
                norm_name = normalize_name_for_dedupe(item_payload.name)
                existing = (
                    db.query(Item)
                    .filter(Item.warehouse_id == item_payload.warehouse_id)
                    .filter(func.lower(func.trim(Item.name)) == norm_name)
                    .filter(func.lower(Item.unit) == item_payload.unit.strip().lower())
                    .first()
                )

            if existing:
                skipped_existing += 1
                continue

            dedupe_key = (item_payload.warehouse_id, normalize_name_for_dedupe(item_payload.name), item_payload.unit.strip().lower())
            if item_payload.product_code is not None:
                dedupe_code = item_payload.product_code.lower()
                if dedupe_code in existing_code_set:
                    skipped_existing += 1
                    continue
                existing_code_set.add(dedupe_code)

            if dedupe_key in existing_name_unit_set:
                skipped_existing += 1
                continue

            existing_name_unit_set.add(dedupe_key)

            created += 1
            if not payload.dry_run:
                db.add(
                    Item(
                        product_code=item_payload.product_code,
                        name=item_payload.name,
                        unit=item_payload.unit,
                        step=item_payload.step,
                        min_qty=item_payload.min_qty,
                        max_qty=item_payload.max_qty,
                        is_favorite=item_payload.is_favorite,
                        is_active=row.is_active if row.is_active is not None else True,
                        warehouse_id=item_payload.warehouse_id,
                        category_id=item_payload.category_id,
                        station_id=item_payload.station_id,
                    )
                )
            else:
                updated += 0
        except Exception as exc:
            errors.append({"row": row_index, "message": str(exc)})

    if not payload.dry_run and not errors:
        db.commit()
    else:
        db.rollback()

    return {
        "dry_run": payload.dry_run,
        "total": len(payload.rows),
        "created": created,
        "updated": updated,
        "skipped_existing": skipped_existing,
        "errors": errors,
    }


@router.get("/categories", response_model=list[ItemCategoryOut])
def list_item_categories(
    warehouse_id: int | None = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    if warehouse_id is None:
        return db.query(ItemCategory).order_by(ItemCategory.name.asc()).all()

    _ensure_warehouse_exists(db, warehouse_id)

    usage_score = func.coalesce(func.sum(ItemUsageStat.use_count), 0).label("usage_score")

    return (
        db.query(ItemCategory)
        .outerjoin(
            Item,
            and_(
                Item.category_id == ItemCategory.id,
                Item.warehouse_id == warehouse_id,
            ),
        )
        .outerjoin(
            ItemUsageStat,
            and_(
                ItemUsageStat.item_id == Item.id,
                ItemUsageStat.warehouse_id == warehouse_id,
            ),
        )
        .group_by(ItemCategory.id)
        .order_by(usage_score.desc(), ItemCategory.name.asc())
        .all()
    )


@router.post("/categories", response_model=ItemCategoryOut, status_code=201)
def create_item_category(
    payload: ItemCategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    category = ItemCategory(name=payload.name)
    db.add(category)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Category already exists")
    db.refresh(category)
    return category


@router.post("/{item_id}/aliases", response_model=ItemAliasOut, status_code=201)
def add_item_alias(
    item_id: int,
    payload: ItemAliasCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    item = db.query(Item.id).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    alias = ItemAlias(item_id=item_id, alias_text=payload.alias_text)
    db.add(alias)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Alias already exists")
    db.refresh(alias)
    return alias


@router.delete("/{item_id}/aliases/{alias_id}", status_code=204)
def delete_item_alias(
    item_id: int,
    alias_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_catalog_manage_role(current_user)

    alias = (
        db.query(ItemAlias).filter(ItemAlias.id == alias_id, ItemAlias.item_id == item_id).first()
    )
    if not alias:
        raise HTTPException(status_code=404, detail="Alias not found")

    db.delete(alias)
    db.commit()
    return None
